import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import os

class ExcelHandler:
    def __init__(self, filename='job_applications.xlsx'):
        self.filename = filename
        self.workbook = None
        self.sheet = None
        self.create_or_load_workbook()

    def create_or_load_workbook(self):
        root_directory = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        excel_file_path = os.path.join(root_directory, self.filename)
        if os.path.exists(excel_file_path):
            self.workbook = openpyxl.load_workbook(excel_file_path)
            self.sheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(["Date", "Job Link", "Job Description", "Cover Letter"])




    def add_application(self, link, description, cover_letter):
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.append([current_date, link, description, cover_letter])
        

    def save_excel(self):
        self.workbook.save(filename=self.filename)

